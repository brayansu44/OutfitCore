from django.shortcuts import render, redirect, get_object_or_404
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.db.models import F
from .models import *
from usuarios.models import PerfilUsuario
from .forms import SalidaProductoForm, StockForm, EntregaCorteForm, InsumoForm
import json
import openpyxl
import os
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, Image
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet

# Vista del inventario de bodega
@login_required(login_url='login')
def inventario_bodega(request):
    stock_queryset = Stock.objects.select_related('producto_variante__producto', 'producto_variante__color', 'producto_variante__talla', 'producto_variante__diseno')

    stock_list = []
    for item in stock_queryset:
        entradas = EntregaCorte.objects.filter(producto=item.producto_variante).exists()
        salidas = SalidaProducto.objects.filter(producto__producto_variante=item.producto_variante).exists()
        tiene_movimientos = entradas or salidas

        stock_list.append({
            'stock': item,
            'tiene_movimientos': tiene_movimientos,
        })

    return render(request, 'bodega/inventario-bodega.html', {'stock': stock_list})

def crear_stock(request):
    if request.method == 'POST':
        form = StockForm(request.POST)
        if form.is_valid():
            producto_variante = form.cleaned_data['producto_variante']
            cantidad = form.cleaned_data['cantidad']

            # Validar si ya existe
            if Stock.objects.filter(producto_variante=producto_variante).exists():
                messages.error(request, "Ya existe este producto en bodega.")
            else:
                form.save()
                messages.success(request, "Producto agregado al inventario de bodega.")
                return redirect('inventario_bodega')
    else:
        form = StockForm()

    return render(request, 'bodega/form_stock.html', {
        'form': form,
        'accion': 'Agregar al inventario'
    })

def editar_stock(request, stock_id):
    stock = get_object_or_404(Stock, id=stock_id)

    if request.method == 'POST':
        form = StockForm(request.POST, instance=stock)
        if form.is_valid():
            form.save()
            messages.success(request, "Inventario actualizado.")
            return redirect('inventario_bodega')
    else:
        form = StockForm(instance=stock)

    return render(request, 'bodega/form_stock.html', {
        'form': form,
        'accion': 'Editar inventario'
    })

@login_required(login_url='login')
@require_POST
def eliminar_stock(request, stock_id):
    if request.method == "POST":
        stock = get_object_or_404(Stock, id=stock_id)
        stock.delete()
        return JsonResponse({"success": True})

    return JsonResponse({"success": False, "error": "Método no permitido"})

# Vista de entradas de productos
@login_required(login_url='login')
def entradas_producto(request):
    entrega_corte = EntregaCorte.objects.all()
    return render(request, 'bodega/entradas-productos.html', {'entrega_corte': entrega_corte})

@login_required
def crear_entrada(request):
    if request.method == 'POST':
        form = EntregaCorteForm(request.POST)
        if form.is_valid():
            entrada = form.save(commit=False)
            try:
                entrada.user_responsable = request.user.perfilusuario
            except PerfilUsuario.DoesNotExist:
                messages.error(request, "Este usuario no tiene un perfil asociado.")
                return redirect('entradas')
            entrada.save()
            messages.success(request, "Entrada registrada exitosamente.")
            return redirect('entradas')
    else:
        form = EntregaCorteForm()
    return render(request, 'bodega/form_entrada.html', {
        'form': form,
        'accion': 'Registrar'
    })

@login_required
def editar_entrada(request, entrada_id):
    entrada = get_object_or_404(EntregaCorte, id=entrada_id)
    if request.method == 'POST':
        form = EntregaCorteForm(request.POST, instance=entrada)
        if form.is_valid():
            form.save()
            messages.success(request, "Entrada actualizada correctamente.")
            return redirect('entradas')
    else:
        form = EntregaCorteForm(instance=entrada)
    return render(request, 'bodega/form_entrada.html', {
        'form': form,
        'accion': 'Editar'
    })

@login_required
def eliminar_entrada(request, entrada_id):
    entrada = get_object_or_404(EntregaCorte, id=entrada_id)
    devolver = request.GET.get('devolver') == '1'

    with transaction.atomic():
        cantidad_en_bodega = entrada.cantidad - entrada.cantidad_lavado

        try:
            stock = Stock.objects.select_for_update().get(producto_variante=entrada.producto)

            if devolver:
                stock.cantidad = F('cantidad') - cantidad_en_bodega
                stock.save()
        except Stock.DoesNotExist:
            pass

        if entrada.cantidad_lavado > 0:
            salida_lavado = SalidaProducto.objects.filter(
                producto__producto_variante=entrada.producto,
                cantidad=entrada.cantidad_lavado,
                es_lavado=True,
                user_responsable=entrada.user_responsable
            ).first()
            if salida_lavado:
                salida_lavado.delete()

        entrada.delete()

    return JsonResponse({'success': True})

# Vista de salidas de productos
@login_required(login_url='login')
def salidas_producto(request):
    salida_producto = SalidaProducto.objects.all() 
    return render(request, 'bodega/salidas-productos.html', {'salida_producto': salida_producto})

@login_required(login_url='login')
def crear_salida_producto(request):
    if request.method == 'POST':
        form = SalidaProductoForm(request.POST, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, 'Salida registrada correctamente.')
            return redirect('salidas_producto')
    else:
        form = SalidaProductoForm(request=request)

    return render(request, 'bodega/form_salida.html', {
        'form': form,
        'accion': 'Registrar salida'
    })

@login_required(login_url='login')
def editar_salida_producto(request, salida_id):
    salida = get_object_or_404(SalidaProducto, id=salida_id)

    if request.method == 'POST':
        form = SalidaProductoForm(request.POST, instance=salida, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, 'Salida actualizada correctamente.')
            return redirect('salidas_producto')
    else:
        form = SalidaProductoForm(instance=salida, request=request)

    return render(request, 'bodega/form_salida.html', {
        'form': form,
        'accion': 'Editar salida'
    })

@csrf_exempt
@login_required(login_url='login')
@require_POST
def eliminar_salida_producto(request, salida_id):
    try:
        salida = get_object_or_404(SalidaProducto, id=salida_id)

        data = json.loads(request.body)
        devolver = data.get("devolver", False)

        if devolver and not salida.es_lavado:
            # Devolver al stock
            stock = salida.producto
            stock.cantidad += salida.cantidad
            stock.save()

        salida.delete()
        return JsonResponse({"success": True})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})
    
def historial_movimientos_producto(request, variante_id):
    variante = get_object_or_404(ProductoVariante, id=variante_id)
    entradas = EntregaCorte.objects.filter(producto=variante)
    salidas = SalidaProducto.objects.filter(producto__producto_variante=variante)

    context = {
        'variante': variante,
        'entradas': entradas,
        'salidas': salidas,
    }
    return render(request, 'bodega/historial_movimientos_producto.html', context)

@login_required
def exportar_historial_excel(request, variante_id):
    variante = get_object_or_404(ProductoVariante, id=variante_id)
    entradas = EntregaCorte.objects.filter(producto=variante).order_by('-fecha')
    salidas = SalidaProducto.objects.filter(producto__producto_variante=variante).order_by('-fecha')

    if not entradas.exists() and not salidas.exists():
        return render(request, 'bodega/sin_movimientos.html', status=400)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial"

    # Información del producto
    ws.append(["Producto:", str(variante.producto.referencia)])
    ws.append(["Color:", str(variante.color)])
    ws.append(["Talla:", str(variante.talla)])
    ws.append(["Diseño:", str(variante.diseno.nombre) if variante.diseno else "Sin diseño"])
    ws.append([])  # Línea en blanco

    # Encabezados
    ws.append(["Tipo", "Fecha", "Cantidad", "Responsable", "Estado"])

    for entrada in entradas:
        ws.append([
            "Entrada",
            entrada.fecha.strftime('%Y-%m-%d'),
            entrada.cantidad,
            str(entrada.user_responsable),
            "Completado"
        ])

    for salida in salidas:
        ws.append([
            "Lavado" if salida.es_lavado else "Salida",
            salida.fecha.strftime('%Y-%m-%d'),
            salida.cantidad,
            str(salida.user_responsable),
            salida.estado
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # Respuesta HTTP
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"Historial_{variante_id}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
def exportar_historial_pdf(request, variante_id):
    variante = get_object_or_404(ProductoVariante, id=variante_id)
    entradas = EntregaCorte.objects.filter(producto=variante).order_by('-fecha')
    salidas = SalidaProducto.objects.filter(producto__producto_variante=variante).order_by('-fecha')

    if not entradas.exists() and not salidas.exists():
        return render(request, 'bodega/sin_movimientos.html', status=400)

    # Configurar respuesta PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Historial_{variante_id}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    # --- Logo ---
    logo_path = os.path.join(settings.BASE_DIR, 'kiddes', 'static', 'images', 'logo-img.png')
    if os.path.exists(logo_path):
        logo = Image(logo_path, width=2 * inch, height=1 * inch)
        elements.append(logo)

    elements.append(Spacer(1, 12))

    # --- Título ---
    title = Paragraph(f"<strong>Historial de Movimientos</strong>", styles['Title'])
    elements.append(title)

    # --- Info del producto ---
    producto_info = f"""
        <strong>Referencia:</strong> {variante.producto.referencia}<br/>
        <strong>Color:</strong> {variante.color}<br/>
        <strong>Talla:</strong> {variante.talla}<br/><br/>
    """
    elements.append(Paragraph(producto_info, styles['Normal']))

    # --- Tabla de movimientos ---
    data = [['Tipo', 'Fecha', 'Cantidad', 'Responsable', 'Estado']]

    for entrada in entradas:
        data.append([
            'Entrada',
            entrada.fecha.strftime('%Y-%m-%d'),
            entrada.cantidad,
            str(entrada.user_responsable),
            'Completado'
        ])

    for salida in salidas:
        tipo = 'Lavado' if salida.es_lavado else 'Salida'
        data.append([
            tipo,
            salida.fecha.strftime('%Y-%m-%d'),
            salida.cantidad,
            str(salida.user_responsable),
            salida.estado
        ])

    if len(data) == 1:
        elements.append(Paragraph("No hay movimientos registrados para este producto.", styles['Normal']))
    else:
        table = Table(data, colWidths=[1.2*inch]*5)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(table)

    # --- Generar PDF ---
    doc.build(elements)
    return response

def inventario_insumos(request):
    insumos = Insumo.objects.all().order_by('nombre')
    return render(request, 'bodega/insumos/inventario_insumos.html', {'insumos': insumos}) 

def agregar_insumo(request):
    if request.method == 'POST':
        form = InsumoForm(request.POST)
        if form.is_valid():
            nombre = form.cleaned_data['nombre']
            if Insumo.objects.filter(nombre__iexact=nombre).exists():
                messages.error(request, 'Ya existe un insumo con ese nombre.')
            else:
                form.save()
                messages.success(request, 'Insumo creado exitosamente.')
                return redirect('inventario_insumos')
    else:
        form = InsumoForm()
    
    return render(request, 'bodega/insumos/form_insumo.html', {
        'form': form,
        'accion': 'Agregar insumo'
    })

def editar_insumo(request, insumo_id):
    insumo = get_object_or_404(Insumo, id=insumo_id)
    
    if request.method == 'POST':
        form = InsumoForm(request.POST, instance=insumo)
        if form.is_valid():
            form.save()
            messages.success(request, 'Insumo actualizado correctamente.')
            return redirect('inventario_insumos')
    else:
        form = InsumoForm(instance=insumo)
    
    return render(request, 'bodega/insumos/form_insumo.html', {
        'form': form,
        'accion': 'Editar insumo'
    })

@login_required
def eliminar_insumo(request, insumo_id):
    insumo = get_object_or_404(Insumo, id=insumo_id)

    with transaction.atomic():
        insumo.delete()

    return JsonResponse({'success': True})
